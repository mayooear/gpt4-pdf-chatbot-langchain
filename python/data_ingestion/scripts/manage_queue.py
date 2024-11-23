#!/usr/bin/env python
"""
Media Ingestion Queue Manager

A robust CLI tool for managing distributed media ingestion workflows. Handles queuing and status 
tracking for audio files and YouTube content across multiple processing environments.

Technical Specifications:
- Queue Backend: DynamoDB for distributed state management
- Storage: AWS S3 with configurable bucket paths
- Processing: Parallel queue support via isolated queue directories
- Monitoring: Real-time status tracking and ETA calculations

Supported Media Types:
- Audio: MP3, WAV, FLAC (with metadata preservation)
- Video: YouTube single videos and playlists
- Bulk Imports: Directory scanning with duplicate detection
- Structured Imports: Excel-based playlist processing
- Text File Imports: List of YouTube URLs (one per line)

Key Operations:
add:
  - Single media (--video URL, --audio path)
  - Bulk import (--directory path)
  - Playlist import (--playlist URL, --playlists-file path)
  - URL list import (--urls-file path)

manage:
  - List queue (--list)
  - Clear queue (--clear)
  - Reset failed (--reset-failed)
  - Remove completed (--remove-completed)
  - Process status (--status)

Configuration:
  Required:
  - library_config.json: Defines target libraries and S3 paths
  - --site: Environment identifier (dev/staging/prod)
  
  Optional:
  - --queue: Alternate queue name for parallel processing
  - --debug: Enhanced logging output

Example Usage:
  Add content:
    ./manage_queue.py --video URL --default-author "Name" --library "LibraryName" --site dev
    ./manage_queue.py --directory /path/to/files --default-author "Name" --library "LibraryName" --site dev
    ./manage_queue.py --urls-file youtube_urls.txt --default-author "Name" --library "LibraryName" --site dev

  Monitor:
    ./manage_queue.py --list --site dev
    ./manage_queue.py --status --site dev

Dependencies:
  AWS: S3, DynamoDB
  APIs: YouTube Data API v3
  Python: >=3.7
"""

import argparse
from datetime import timedelta
import os
import logging
from tqdm import tqdm
from openpyxl import load_workbook
from collections import defaultdict
import pytz
from datetime import datetime
import json
from data_ingestion.scripts.IngestQueue import IngestQueue
from data_ingestion.scripts.logging_utils import configure_logging
from data_ingestion.scripts.youtube_utils import (
    extract_youtube_id, 
    get_playlist_videos,
    load_youtube_data_map
)
from data_ingestion.scripts.media_utils import get_file_hash
from data_ingestion.scripts.processing_time_estimates import get_estimate, estimate_total_processing_time
from util.env_utils import load_env

# Load library configuration
with open('library_config.json', 'r') as f:
    LIBRARY_CONFIG = json.load(f)

logger = logging.getLogger(__name__)


def initialize_environment(args):
    """
    Sets up the runtime environment based on deployment site (dev/staging/prod).
    Configures logging verbosity and loads appropriate AWS credentials.
    """
    load_env(args.site)
    configure_logging(args.debug)


def get_unique_files(directory_path):
    """
    Recursively scans directory for media files and deduplicates using file hashes.
    
    Performance Note: For large directories, this performs I/O-intensive hash calculations.
    Memory Usage: Stores only file paths and hashes, not file contents.
    
    Returns: List of unique file paths, preserving the first occurrence of duplicate content
    """
    unique_files = {}
    files_to_check = []

    # Collect all relevant files
    for root, _, files in os.walk(directory_path):
        for file in files:
            if file.lower().endswith((".mp3", ".wav", ".flac", ".mp4", ".avi", ".mov")):
                files_to_check.append(os.path.join(root, file))

    # Process files with progress bar for long-running operations
    for file_path in tqdm(files_to_check, desc="Checking for unique files", ncols=100):
        file_hash = get_file_hash(file_path)
        if file_hash not in unique_files:
            unique_files[file_hash] = file_path

    return list(unique_files.values())


def process_audio_input(input_path, queue, default_author, library):
    """
    Routes audio processing based on input type (single file vs directory).
    
    Validation:
    - Verifies library exists in config
    - Checks file extensions
    - Ensures input path exists
    
    S3 Path Structure: public/audio/{library_name}/{relative_path}
    """
    # Check if the library is in the config file
    if library not in LIBRARY_CONFIG:
        error_msg = f"Error: Library '{library}' not found in library_config.json. Please use a valid library name."
        logger.error(error_msg)
        raise ValueError(error_msg)

    if os.path.isfile(input_path):
        if input_path.lower().endswith((".mp3", ".wav", ".flac")):
            s3_folder = library.lower()  
            s3_key = f"public/audio/{s3_folder}/{os.path.basename(input_path)}"
            item_id = queue.add_item(
                "audio_file",
                {
                    "file_path": input_path,
                    "author": default_author,
                    "library": LIBRARY_CONFIG[library],
                    "s3_folder": s3_folder,
                    "s3_key": s3_key
                },
            )
            if item_id:
                logger.info(f"Added audio file to queue: {item_id} - {input_path}")
                return [item_id]
            else:
                logger.error(f"Failed to add audio file to queue: {input_path}")
                return []
        else:
            logger.error(f"Unsupported file type: {input_path}")
            return []
    elif os.path.isdir(input_path):
        return process_directory(input_path, queue, default_author, library)
    else:
        logger.error(f"Invalid input path: {input_path}")
        return []


def process_directory(directory_path, queue, default_author, library):
    # Check if the library is in the config file
    if library not in LIBRARY_CONFIG:
        error_msg = f"Error: Library '{library}' not found in library_config.json. Please use a valid library name."
        logger.error(error_msg)
        raise ValueError(error_msg)

    unique_files = get_unique_files(directory_path)
    added_items = []
    s3_folder = library.lower() 

    # Process unique files with tqdm
    for file_path in tqdm(unique_files, desc="Processing unique files"):
        # Calculate the relative path
        relative_path = os.path.relpath(file_path, directory_path)
        # Combine the s3_folder with the relative path
        s3_key = f"public/audio/{s3_folder}/{relative_path}"
        
        item_id = queue.add_item(
            "audio_file",
            {
                "file_path": file_path,
                "author": default_author,
                "library": LIBRARY_CONFIG[library],
                "s3_folder": s3_folder,
                "s3_key": s3_key
            },
        )
        if item_id:
            added_items.append(item_id)
        else:
            logger.error(f"Failed to add audio file to queue: {file_path}")

    return added_items


def add_to_queue(args, queue, source=None):
    """
    Central routing function for all queue additions. Handles YouTube videos, playlists, and audio files.
    
    source: Optional playlist URL for tracking video origins in bulk imports
    """
    if args.video:
        youtube_id = extract_youtube_id(args.video)
        if youtube_id:
            item_id = queue.add_item(
                "youtube_video",
                {
                    "url": args.video,
                    "youtube_id": youtube_id,
                    "author": args.default_author,
                    "library": args.library,
                    "source": source,  # Tracks origin playlist for duplicate detection
                },
            )
            if item_id:
                logger.info(f"Added YouTube video to queue: {item_id}")
            else:
                logger.error("Failed to add YouTube video to queue")
        else:
            logger.error("Invalid YouTube video URL")

    elif args.playlist:
        # Bulk process all videos in playlist while preserving order
        videos = get_playlist_videos(args.playlist)
        for video in videos:
            logger.debug(f"Video to add: {video}")
            item_id = queue.add_item(
                "youtube_video",
                {
                    "url": video["url"],
                    "youtube_id": video["youtube_id"],
                    "author": args.default_author,
                    "library": args.library,
                    "source": args.playlist,
                },
            )
            if item_id:
                logger.info(f"Added YouTube video from playlist to queue: {item_id}")
            else:
                logger.error(f"Failed to add YouTube video to queue: {video['url']}")

    elif args.audio or args.directory:
        input_path = args.audio or args.directory
        added_items = process_audio_input(input_path, queue, args.default_author, args.library)
        if added_items:
            logger.info(f"Added {len(added_items)} audio file(s) to queue")
        else:
            logger.error(f"Failed to add any audio files from: {input_path}")

    else:
        logger.error("No valid input provided for adding to queue")


def truncate_path(file_path, num_dirs=3):
    """
    Shortens file paths for display while preserving meaningful context.
    Example: /very/long/path/to/file.mp3 -> .../path/to/file.mp3
    
    num_dirs: Number of trailing directories to preserve
    """
    parts = file_path.split(os.sep)
    return os.sep.join(parts[-num_dirs - 1 :]) if len(parts) > num_dirs else file_path


def list_queue_items(queue):
    """
    Displays queue contents in a formatted table with dynamic column widths.
    Converts timestamps to PST and truncates long paths for readability.
    
    Output format:
    ID  Type  Status  URL/File  Last Updated
    """
    items = queue.get_all_items()
    if not items:
        print("The queue is empty.")
        return

    print("Queue contents:")
    # Determine the maximum lengths for each column dynamically
    max_id_len = max(
        (len(str(item.get("id", "Unknown ID"))) for item in items), default=10
    )
    max_type_len = max(
        (len(item.get("type", "Unknown type")) for item in items), default=15
    )
    max_status_len = max(
        (len(item.get("status", "Unknown status")) for item in items), default=15
    )
    max_url_len = max(
        (
            len(item.get("data", {}).get("url", ""))
            for item in items
            if item.get("type") == "youtube_video"
        ),
        default=50,
    )
    max_file_path_len = max(
        (
            len(truncate_path(item.get("data", {}).get("file_path", "")))
            for item in items
            if item.get("type") == "audio_file"
        ),
        default=50,
    )
    max_updated_len = max(
        (len(item.get("last_updated", "Unknown time")) for item in items), default=20
    )

    # Print header with consistent spacing
    print(
        f"{'ID'.ljust(max_id_len)}  {'Type'.ljust(max_type_len)}  {'Status'.ljust(max_status_len)}  {'URL/File'.ljust(max(max_url_len, max_file_path_len))}  {'Last Updated'.ljust(max_updated_len)}"
    )

    for item in items:
        item_id = str(item.get("id", "Unknown ID"))
        item_type = item.get("type", "Unknown type")
        item_data = item.get("data", {})
        item_status = item.get("status", "Unknown status")
        last_updated = item.get("updated_at", "Unknown time")

        # Convert last_updated to PST
        if last_updated != "Unknown time":
            utc_time = datetime.strptime(last_updated, "%Y-%m-%dT%H:%M:%S.%f")
            utc_time = utc_time.replace(tzinfo=pytz.utc)
            pst_time = utc_time.astimezone(pytz.timezone("America/Los_Angeles"))
            last_updated = pst_time.strftime("%Y-%m-%d %H:%M")
            debug_message = f"Last updated: {last_updated}"
            logger.debug(debug_message)

        # Format output based on item type
        if item_type == "youtube_video":
            print(
                f"{item_id.ljust(max_id_len)}  {item_type.ljust(max_type_len)}  {item_status.ljust(max_status_len)}  {item_data.get('url', '').ljust(max(max_url_len, max_file_path_len))}  {last_updated.ljust(max_updated_len)}"
            )
        elif item_type == "audio_file":
            truncated_path = truncate_path(item_data.get("file_path", ""))
            print(
                f"{item_id.ljust(max_id_len)}  {item_type.ljust(max_type_len)}  {item_status.ljust(max_status_len)}  {truncated_path.ljust(max(max_url_len, max_file_path_len))}  {last_updated.ljust(max_updated_len)}"
            )
        else:
            print(
                f"{item_id.ljust(max_id_len)}  {item_type.ljust(max_type_len)}  {item_status.ljust(max_status_len)}  {'N/A'.ljust(max(max_url_len, max_file_path_len))}  {last_updated.ljust(max_updated_len)}"
            )
    print_queue_status(queue, items)


def clear_queue(queue):
    queue.clear_queue()
    logger.info("Queue has been cleared")


def reset_stuck_items(queue):
    reset_count = queue.reset_stuck_items()
    logger.info(
        f"Reset {reset_count} items from error or interrupted state. Ready for processing."
    )

def remove_completed_items(queue):
    removed_count = queue.remove_completed_items()
    logger.info(f"Removed {removed_count} completed items from the queue")

def reprocess_item(queue, item_id):
    success, message = queue.reprocess_item(item_id)
    if success:
        logger.info(message)
    else:
        logger.warning(message)


def reprocess_all_items(queue):
    reset_count = queue.reset_all_items()
    logger.info(f"Reset all {reset_count} items in the queue. Ready for reprocessing.")


def reset_processing_items(queue):
    reset_count = queue.reset_processing_items()
    logger.info(
        f"Reset {reset_count} items from processing state to pending. Ready for processing."
    )


def remove_item(queue, item_id):
    if queue.remove_item(item_id):
        logger.info(f"Successfully removed item {item_id} from the queue")
    else:
        logger.error(f"Failed to remove item {item_id} from the queue")


def process_playlists_file(args, queue):
    """
    Processes an Excel file containing playlist information for bulk imports.
    
    Expected Excel format:
    Title | Default Author | Library | Playlist URL
    
    Features:
    - Deduplicates videos across playlists
    - Tracks source playlists for duplicate videos
    - Maintains original playlist metadata
    """
    workbook = load_workbook(filename=args.playlists_file, read_only=True)
    sheet = workbook.active
    
    all_videos = []
    video_sources = defaultdict(list)  # Tracks which playlists contain each video
    processed_playlists = 0

    # Skip header row and process each playlist
    for row in sheet.iter_rows(min_row=2, values_only=True):
        title, default_author, library, playlist_url = row
        videos = get_playlist_videos(playlist_url)
        processed_playlists += 1
        
        # Track video sources for duplicate reporting
        for video in videos:
            all_videos.append(video)
            video_sources[video['url']].append(title)

    # Deduplicate videos while preserving first occurrence
    unique_videos = {v['url']: v for v in all_videos}.values()
    duplicates_removed = len(all_videos) - len(unique_videos)

    # Add unique videos to queue with original metadata
    for video in unique_videos:
        args.video = video['url']
        args.default_author = default_author
        args.library = library
        add_to_queue(args, queue, source=playlist_url)

    # Log processing summary and duplicates
    logger.info(f"Processed {processed_playlists} playlists")
    logger.info(f"Total videos found: {len(all_videos)}")
    logger.info(f"Unique videos added to queue: {len(unique_videos)}")

    if duplicates_removed > 0:
        logger.info(f"{duplicates_removed} duplicate videos (not added to queue):")
        for url, sources in video_sources.items():
            if len(sources) > 1:
                logger.info(f"{url} (Found in: {', '.join(sources)})")


def print_queue_status(queue, items=None):
    if items is None:
        items = queue.get_all_items()

    estimated_time = estimate_total_processing_time(items)
    print(f"Estimated time to complete all pending items: {estimated_time}")

    audio_estimate = get_estimate("audio_file")
    video_estimate = get_estimate("youtube_video")
    
    if audio_estimate and audio_estimate.get('time') is not None and audio_estimate.get('size') is not None:
        print(f"Average processing time for audio files: {timedelta(seconds=int(audio_estimate['time']))} for {audio_estimate['size'] / (1024*1024):.2f} MB")
    else:
        print("No data available for audio file processing times.")
    
    if video_estimate and video_estimate.get('time') is not None and video_estimate.get('size') is not None:
        print(f"Average processing time for YouTube videos: {timedelta(seconds=int(video_estimate['time']))} for {video_estimate['size'] / (1024*1024):.2f} MB")
    else:
        print("No data available for YouTube video processing times.")


def process_urls_file(args, queue):
    """
    Process a text file containing YouTube URLs (one per line).
    Skips duplicates and already processed videos.
    """
    if not os.path.exists(args.urls_file):
        logger.error(f"URLs file not found: {args.urls_file}")
        logger.error("Please provide the correct path to your YouTube URLs file")
        logger.error("Example file format (one URL per line):")
        logger.error("https://youtu.be/VIDEO_ID1")
        logger.error("https://youtu.be/VIDEO_ID2")
        return
    
    with open(args.urls_file, 'r') as f:
        urls = [line.strip() for line in f if line.strip()]
    
    if not urls:
        logger.error(f"No valid URLs found in {args.urls_file}")
        return
        
    processed = 0
    skipped = 0
    
    for url in urls:
        youtube_id = extract_youtube_id(url)
        if not youtube_id:
            logger.error(f"Invalid YouTube URL: {url}")
            continue
            
        # Check if already processed via youtube_data_map
        youtube_data_map = load_youtube_data_map()
        if youtube_id in youtube_data_map:
            logger.info(f"Skipping already processed video: {url}")
            skipped += 1
            continue
            
        item_id = queue.add_item(
            "youtube_video",
            {
                "url": url,
                "youtube_id": youtube_id,
                "author": args.default_author,
                "library": args.library,
            },
        )
        if item_id:
            processed += 1
            logger.info(f"Added YouTube video to queue: {item_id} - {url}")
        else:
            logger.error(f"Failed to add YouTube video to queue: {url}")
    
    logger.info(f"Processed {processed} videos")
    logger.info(f"Skipped {skipped} already processed videos")


def main():
    """
    Main CLI entry point with comprehensive argument parsing and operation routing.
    
    Operation Groups:
    1. Queue Management (clear, reset, remove)
    2. Content Addition (video, audio, playlists)
    3. Status/Monitoring (list, status)
    
    Environment Setup:
    - Requires site parameter for AWS credential configuration
    - Optional debug mode for enhanced logging
    - Supports parallel queues via --queue parameter
    """
    parser = argparse.ArgumentParser(description="Manage the ingest queue")

    # Add operation arguments
    parser.add_argument("--video", help="YouTube video URL")
    parser.add_argument("--playlist", help="YouTube playlist URL")
    parser.add_argument("--audio", help="Path to audio file")
    parser.add_argument("--directory", help="Path to directory containing audio files")
    parser.add_argument("--default-author", help="Default author of the media")
    parser.add_argument("--library", help="Name of the library")
    parser.add_argument(
        "--list", action="store_true", help="List all items in the processing queue"
    )
    parser.add_argument(
        "--clear", action="store_true", help="Clear all items from the queue"
    )
    parser.add_argument(
        "--reset-failed",
        action="store_true", 
        help="Reset items in error or interrupted state back to pending"
    )
    parser.add_argument(
        "--remove-completed",
        action="store_true",
        help="Remove all completed items from the queue",
    )
    parser.add_argument("--debug", action="store_true", help="Enable debug logging")
    parser.add_argument("--reprocess", help="Reprocess a specific item by ID")
    parser.add_argument(
        "--reprocess-all",
        action="store_true",
        help="Reset all items in the queue for reprocessing",
    )
    parser.add_argument("--playlists-file", help="Path to XLSX file containing playlist information")
    parser.add_argument("--queue", default=None, help="Specify an alternative queue name")
    parser.add_argument(
        "--reset-processing-items",
        action="store_true",
        help="Reset items in processing state to pending",
    )
    parser.add_argument("--remove", help="Remove a specific item from the queue by ID")
    parser.add_argument("--status", action="store_true", help="Print the queue status")
    parser.add_argument('--site', required=True, help='Site ID for environment variables')
    parser.add_argument(
        "--urls-file", 
        help="Path to text file containing YouTube URLs (one per line)"
    )
    args = parser.parse_args()

    # Initialize environment and create queue instance
    initialize_environment(args)
    queue = IngestQueue(queue_dir=args.queue) if args.queue else IngestQueue()

    if args.queue:
        logger.info(f"Using queue: {args.queue}")

    # Route to appropriate operation handler
    if args.status:
        print_queue_status(queue)
    elif args.remove:
        remove_item(queue, args.remove)
    elif args.playlists_file:
        process_playlists_file(args, queue)
    elif args.reprocess_all:
        reprocess_all_items(queue)
    elif args.reprocess:
        reprocess_item(queue, args.reprocess)
    elif args.list:
        list_queue_items(queue)
    elif args.clear:
        clear_queue(queue)
    elif args.reset_failed:
        reset_stuck_items(queue)
    elif args.reset_processing_items:
        reset_processing_items(queue)
    elif args.remove_completed:
        remove_completed_items(queue)
    elif args.urls_file:
        if not args.default_author or not args.library:
            logger.error("For adding items, you must specify both --default-author and --library")
            parser.print_help()
            return
        process_urls_file(args, queue)
    elif any([args.video, args.playlist, args.audio, args.directory]):
        if not args.default_author or not args.library:
            logger.error("For adding items, you must specify both --default-author and --library")
            parser.print_help()
            return
        add_to_queue(args, queue)
    else:
        logger.error("No valid operation specified.")
        parser.print_help()
        return

    # Display final queue status
    queue_status = queue.get_queue_status()
    print(f"Queue status: {queue_status}")


if __name__ == "__main__":
    main()
